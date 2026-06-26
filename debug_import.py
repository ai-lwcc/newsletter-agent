from openpyxl import load_workbook
from core.models import Person

workbook = load_workbook("client profiles-1.xlsx")
sheet = workbook.active

headers = [cell.value for cell in sheet[1]]

field_max_lengths = {
    field.name: field.max_length
    for field in Person._meta.fields
    if getattr(field, "max_length", None)
}

column_to_field = {
    "First Name": "first_name",
    "Last Name": "last_name",
    "Full Name": "full_name",
    "Email": "email",
    "Phone": "phone_number",
    "Phone Number": "phone_number",
    "WhatsApp": "whatsapp_number",
    "WhatsApp Number": "whatsapp_number",
    "Age": "age",
    "Church/Organization": "church_organization",
    "Emergency Contact": "emergency_contact",
    "Emergency Contact Phone": "emergency_contact_phone",
    "Referrer": "referrer",
    "Contact Type": "contact_type",
}

for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    row_data = dict(zip(headers, row))

    for column_name, field_name in column_to_field.items():
        value = row_data.get(column_name)

        if value is None:
            continue

        value = str(value).strip()
        max_length = field_max_lengths.get(field_name)

        if max_length and len(value) > max_length:
            print(
                "Too long:",
                "row", row_number,
                "column", column_name,
                "field", field_name,
                "max", max_length,
                "length", len(value),
                "value:", value,
            )