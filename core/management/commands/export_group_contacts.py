from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from django.core.management.base import BaseCommand

from core.models import Group, Person


class Command(BaseCommand):
    help = "Export the complete LWCC Master Contact List"

    def handle(self, *args, **options):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Master Contacts"

        # Get every group in the system
        groups = list(Group.objects.order_by("name"))

        # Standard contact information
        headers = [
            "First Name",
            "Last Name",
            "Full Name",
            "Email",
            "Phone",
            "WhatsApp",
            "Address",
            "Email Consent",
            "WhatsApp Consent",
            "Active",
            "Subscription Token",
        ]

        # Add one column for every group
        headers.extend(group.name for group in groups)

        sheet.append(headers)

        # Format header row
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Freeze the first row
        sheet.freeze_panes = "A2"

        # Enable filters
        sheet.auto_filter.ref = sheet.dimensions

        people = (
            Person.objects
            .prefetch_related("groups")
            .order_by("last_name", "first_name")
        )

        for person in people:

            person_groups = {
                group.name
                for group in person.groups.all()
            }

            row = [
                person.first_name,
                person.last_name,
                person.full_name,
                person.email,
                person.phone_number,
                person.whatsapp_number,
                person.address,
                "Yes" if person.email_consent else "",
                "Yes" if person.whatsapp_consent else "",
                "Yes" if person.is_active else "",
                str(person.subscription_token),
            ]

            # Automatically populate every group column
            for group in groups:
                row.append(
                    "Yes" if group.name in person_groups else ""
                )

            sheet.append(row)

        # Auto-size every column
        for column_cells in sheet.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(length + 2, 50)

        downloads_folder = Path.home() / "Downloads"
        downloads_folder.mkdir(exist_ok=True)

        filename = downloads_folder / "LWCC Master Contact List.xlsx"

        workbook.save(filename)

        self.stdout.write(
            self.style.SUCCESS(
                f"Created {filename}"
            )
        )