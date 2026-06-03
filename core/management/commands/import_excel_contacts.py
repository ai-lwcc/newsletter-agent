from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from core.contact_importer import import_contact_rows, validate_headers


class Command(BaseCommand):
    help = "Import contacts from an Excel .xlsx file."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"])

        if not excel_path.exists():
            raise CommandError(f"File not found: {excel_path}")

        workbook = load_workbook(excel_path)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        try:
            validate_headers(headers)
        except ValueError as error:
            raise CommandError(str(error))

        rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

        result = import_contact_rows(rows)

        self.stdout.write(
            self.style.SUCCESS(
                f"Excel import complete. Created: {result['created']}, "
                f"Updated: {result['updated']}, Skipped: {result['skipped']}"
            )
        )