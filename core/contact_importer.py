from openpyxl import load_workbook

from core.models import Group, Person


PERSON_FIELD_COLUMNS = {
    "First Name": "first_name",
    "Last Name": "last_name",
    "Full Name": "full_name",
    "Email": "email",
    "Phone": "phone_number",
    "Phone Number": "phone_number",
    "WhatsApp": "whatsapp_number",
    "WhatsApp Number": "whatsapp_number",
    "Age": "age",
    "Address": "address",
    "Church/Organization": "church_organization",
    "Emergency Contact": "emergency_contact",
    "Emergency Contact Phone": "emergency_contact_phone",
    "Referrer": "referrer",
    "Contact Type": "contact_type",
    "Email Consent": "email_consent",
    "WhatsApp Consent": "whatsapp_consent",
    "Notes": "notes",
}

REQUIRED_HEADERS = ["Full Name", "Email"]


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_bool(value):
    return clean_value(value).lower() in [
        "yes",
        "true",
        "1",
        "y",
        "checked",
    ]


def is_valid_email(email):
    email = clean_value(email).lower()
    return "@" in email and "." in email


def validate_headers(headers):
    missing_headers = [
        header for header in REQUIRED_HEADERS if header not in headers
    ]

    if missing_headers:
        raise ValueError(f"Missing columns: {missing_headers}")


def get_dynamic_group_names(row_data):
    group_names = []

    groups_text = clean_value(row_data.get("Groups"))

    if groups_text:
        group_names.extend(
            group.strip()
            for group in groups_text.split(",")
            if group.strip()
        )

    for column_name, value in row_data.items():
        if column_name in PERSON_FIELD_COLUMNS:
            continue

        if column_name == "Groups":
            continue

        if parse_bool(value):
            group_names.append(column_name)

    return sorted(set(group_names))


def build_person_defaults(row_data):
    return {
        "full_name": clean_value(row_data.get("Full Name")),
        "first_name": clean_value(row_data.get("First Name")),
        "last_name": clean_value(row_data.get("Last Name")),
        "phone_number": clean_value(
            row_data.get("Phone") or row_data.get("Phone Number")
        ),
        "whatsapp_number": clean_value(
            row_data.get("WhatsApp") or row_data.get("WhatsApp Number")
        ),
        "age": clean_value(row_data.get("Age")),
        "address": clean_value(row_data.get("Address")),
        "church_organization": clean_value(
            row_data.get("Church/Organization")
        ),
        "emergency_contact": clean_value(
            row_data.get("Emergency Contact")
        ),
        "emergency_contact_phone": clean_value(
            row_data.get("Emergency Contact Phone")
        ),
        "referrer": clean_value(row_data.get("Referrer")),
        "contact_type": clean_value(row_data.get("Contact Type")),
        "notes": clean_value(row_data.get("Notes")),
        "email_consent": parse_bool(row_data.get("Email Consent")),
        "whatsapp_consent": parse_bool(row_data.get("WhatsApp Consent")),
        "is_active": True,
    }


def empty_result():
    return {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "would_create": 0,
        "would_update": 0,
        "would_skip": 0,
        "created_people": [],
        "updated_people": [],
        "skipped_rows": [],
        "groups_to_create": [],
        "groups_created": [],
    }


def preview_contact_rows(rows):
    result = empty_result()
    groups_to_create = set()

    for row_number, row_data in enumerate(rows, start=2):
        email = clean_value(row_data.get("Email")).lower()
        full_name = clean_value(row_data.get("Full Name"))

        if not full_name:
            result["would_skip"] += 1
            result["skipped"] += 1
            result["skipped_rows"].append(
                {"row": row_number, "reason": "Missing full name"}
            )
            continue

        if not is_valid_email(email):
            result["would_skip"] += 1
            result["skipped"] += 1
            result["skipped_rows"].append(
                {"row": row_number, "reason": "Missing or invalid email"}
            )
            continue

        if Person.objects.filter(email=email).exists():
            result["would_update"] += 1
            result["updated_people"].append(full_name)
        else:
            result["would_create"] += 1
            result["created_people"].append(full_name)

        for group_name in get_dynamic_group_names(row_data):
            if not Group.objects.filter(name=group_name).exists():
                groups_to_create.add(group_name)

    result["groups_to_create"] = sorted(groups_to_create)
    return result


def import_contact_rows(rows, dry_run=False):
    if dry_run:
        return preview_contact_rows(rows)

    result = empty_result()
    groups_created = set()

    for row_number, row_data in enumerate(rows, start=2):
        email = clean_value(row_data.get("Email")).lower()
        full_name = clean_value(row_data.get("Full Name"))

        if not full_name:
            result["skipped"] += 1
            result["skipped_rows"].append(
                {"row": row_number, "reason": "Missing full name"}
            )
            continue

        if not is_valid_email(email):
            result["skipped"] += 1
            result["skipped_rows"].append(
                {"row": row_number, "reason": "Missing or invalid email"}
            )
            continue

        defaults = build_person_defaults(row_data)

        person, created = Person.objects.update_or_create(
            email=email,
            defaults=defaults,
        )

        groups = []

        for group_name in get_dynamic_group_names(row_data):
            group, group_created = Group.objects.get_or_create(
                name=group_name,
            )

            if group_created:
                groups_created.add(group_name)

            groups.append(group)

        person.groups.set(groups)

        if created:
            result["created"] += 1
            result["created_people"].append(person.full_name)
        else:
            result["updated"] += 1
            result["updated_people"].append(person.full_name)

    result["groups_created"] = sorted(groups_created)
    return result


def read_excel_contact_rows(excel_path):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    validate_headers(headers)

    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

    return rows